


#https://pygad.readthedocs.io/en/latest/README_pygad_ReadTheDocs.html#cal-pop-fitness

import pygad
import numpy
import matplotlib.pyplot as plt
import time
import numpy as np
import math

from openpyxl import load_workbook
import os
import xlsxwriter

from openpyxl import workbook
from openpyxl import load_workbook





number_of_genes = 10

num_generations=20
sol_per_pop=50



gene_space = [{'low': 0.0000645, 'high': 0.0226, "step": 0.00001},
              {'low': 0.0000645, 'high': 0.0226, "step": 0.00001},
              {'low': 0.0000645, 'high': 0.0226, "step": 0.00001},
              {'low': 0.0000645, 'high': 0.0226, "step": 0.00001},
              {'low': 0.0000645, 'high': 0.0226, "step": 0.00001},
              {'low': 0.0000645, 'high': 0.0226, "step": 0.00001},
              {'low': 0.0000645, 'high': 0.0226, "step": 0.00001},
              {'low': 0.0000645, 'high': 0.0226, "step": 0.00001},
              {'low': 0.0000645, 'high': 0.0226, "step": 0.00001},
              {'low': 0.0000645, 'high': 0.0226, "step": 0.00001}]



recordingAll = []
recAllFit = []


Listx = [None] * (num_generations)



# ------------------------------------------------
# Writing the number of iterations to file
t_file = open('C:/Users/TommyHielscher/Desktop/Benchmark Optimization/10 Bar Truss/Control/number_of_sols.txt','w')
t_file.write(str(num_generations*sol_per_pop)+ " " + "GA")
t_file.close()

# Write termination function
t_file = open('C:/Users/TommyHielscher/Desktop/Benchmark Optimization/10 Bar Truss/Control/termination.txt','w')
t_file.write("False")
t_file.close()
# ------------------------------------------------


#createFolder('C:/Users/TommyHielscher/Desktop/Benchmark Optimization/10 Bar Truss/Optimization_Output/Generation_0/')

def on_start(ga):
    print("This is the start")


    init =  ga_instance.initial_population

    #print(init.shape[0])

    # For each child to be generated
    for _ in range(sol_per_pop):

        candidate = init[_]

        """print("Tests" + str(_))

        # Workbook() takes one, non-optional, argument
        # which is the filename that we want to create.
        filepath = "C:/Users/TommyHielscher/Desktop/Benchmark Optimization/10 Bar Truss/Optimization_Output/Generation_0/" + 'Solution_' + str(_) + '.xlsx'

        workbook = xlsxwriter.Workbook(filepath)

        # The workbook object is then used to add new
        # worksheet via the add_worksheet() method.
        worksheet = workbook.add_worksheet()

        # Use the worksheet object to write
        # data via the write() method.
        worksheet.write('A1', 'Flange Thickness')
        worksheet.write('A2', candidate[0])

        worksheet.write('B1', 'Web Thickness')
        worksheet.write('B2', candidate[1])

        worksheet.write('C1', 'Depth')
        worksheet.write('C2', candidate[2])

        worksheet.write('D1', 'Width')
        worksheet.write('D2', candidate[3])

        worksheet.write('E1', 'Fitness')
        worksheet.write('E2', 'X')


        workbook.close()"""

        """f = open("C:/Users/TommyHielscher/Desktop/Benchmark Optimization/10 Bar Truss/Control/parameters" + str(
            _+1) + ".txt","w")
        f.write(
            str(candidate[0]) + "," + str(candidate[1]) + "," + str(candidate[2]) + "," + str(candidate[3]) + "," + str(candidate[4]) + "," + str(
                candidate[5]) + "," + str(candidate[6]) + "," + str(candidate[7]) + ","
            + str(candidate[8]) + "," + str(candidate[9])
            + "," + "newdata")
        f.close()"""

        print("Generate Initial Pop")

        ##############################################################################




def on_generation(ga):
    print("New Generation")

    print("Generation", ga.generations_completed)
    #print(ga.population)


    fitness_func.counter = 0

    print('Gens complete:' + str(ga.generations_completed) )

    """# Create folder of the next gen and put in the population

    #createFolder('C:/Users/TommyHielscher/Desktop/Benchmark Optimization/10 Bar Truss/Optimization_Output/Generation_' + str(ga.generations_completed) + '/')

    filepath = "C:/Users/TommyHielscher/Desktop/Benchmark Optimization/10 Bar Truss/Control/parameters"

    nextpop = ga_instance.population

    on_generation.counting += 1

    for i in range(sol_per_pop):


        popsols = nextpop[i]

        f = open(filepath + str(ga.generations_completed*sol_per_pop + i) + '.txt',"w")

        f.write(
            str(popsols[0]) + "," + str(popsols[1]) + "," + str(popsols[2]) + "," + str(popsols[3]) + "," + str(popsols[4]) + "," + str(
                popsols[5]) + "," + str(popsols[6]) + "," + str(popsols[7]) + ","
            + str(popsols[8]) + "," + str(popsols[9])
            + "," + "newdata")
        f.close()"""


        ##############################################################################

on_generation.counting = 0



def fitness_func(solution, solution_idx):

    capture = [solution[0], solution[1], solution[2], solution[3], solution[4], solution[5], solution[6], solution[7], solution[8], solution[9]]

    # Write solutions


    f = open("C:/Users/TommyHielscher/Desktop/Benchmark Optimization/10 Bar Truss/Control/parameters" + str(fitness_func.solcounter) + ".txt",
             "w")
    f.write(
        str(solution[0]) + "," + str(solution[1]) + ","+str(solution[2]) + "," + str(solution[3]) + ","+ str(solution[4]) + "," + str(solution[5]) + ","+ str(solution[6]) + ","+ str(solution[7]) + ","
        + str(solution[8]) + ","+ str(solution[9])
        + "," + "newdata")
    f.close()





    #print("Generation:  ", on_generation.counting, " , Solution: ", fitness_func.counter + 1 )

    print('Solution number: ', fitness_func.solcounter )

    # TEST #







    ## READ TJE RESULTS
    filepath = 'C:/Users/TommyHielscher/Desktop/Benchmark Optimization/10 Bar Truss/Output/' + 'abaqus_output_def'+ str( fitness_func.solcounter ) + '.txt'

    while True:
        time.sleep(0.5)

        try:
            file = open(filepath, 'r')
            data = file.read()
            data = data.split(' ')
            data = list(filter(None, data))
            vals = data[-9:-1]
            print("Opening File")
        except:
            print('File has not been created. Waiting for analysis. Current Sol: ' + str(fitness_func.solcounter))

            continue

        break
    print('File found!')

    vals = [float(i) for i in vals]
    print(vals)



    print("Max deflection: " , np.max(vals))
    max_def = np.max(vals)

    # This fitness function value is not correct (we want to minimize the overall weight of the structure)
    diag_l = math.sqrt(9.14**2 + 9.14**2)
    fitness =  (solution[0]*9.14 + solution[1]*9.14 + solution[2]*diag_l + solution[3]*diag_l  + solution[4]*9.14 + solution[5]*diag_l  + solution[6]*diag_l  + solution[7]*9.14 + solution[8]*9.14 + solution[9]*9.14) * 2770

    # I need to rethink how to handle the deflection constraint. This is a crude method to just get it working
    if max_def > 0.0508:
        fitness += 100000*max_def
        fitness = 1 / (fitness+0.000001)

        print("Bad solution: Fitness = ", fitness)
    else:

        fitness = 1 / (fitness + 0.000001)
        print("Fitness =", fitness)

    #print("_____________________")
    #print(X)
    #print("_____________________")

    #A = 10
    #y = A * 2 + sum([(x ** 2 - A * np.cos(2 * math.pi * x)) for x in X])

    #print("Fitness: ", fitness)

    file_ad = "C:/Users/TommyHielscher/Desktop/Benchmark Optimization/10 Bar Truss/Optimization_Output/GA/result_" + str(fitness_func.solcounter) + ".txt"


    t_file = open( file_ad, 'w')
    t_file.write(str(solution) + "," + str(max_def) + "," + str(fitness))

    t_file.close()





    fitness_func.counter += 1

    fitness_func.solcounter += 1

    """"if fitness_func.counter >= sol_per_pop:
         fitness_func.counter = 0"""""




    #################################################



    recordingAll.append(capture)
    recAllFit.append(fitness)

    return fitness


fitness_func.counter = 0
fitness_func.reading = 0

fitness_func.solcounter = 1

def parent_selection_func(fitness, num_parents, ga_instance):

    fitness_sorted = sorted(range(len(fitness)), key=lambda k: fitness[k])
    fitness_sorted.reverse()

    parents = numpy.empty((num_parents, ga_instance.population.shape[1]))

    for parent_num in range(num_parents):
        parents[parent_num, :] = ga_instance.population[fitness_sorted[parent_num], :].copy()

    return parents, fitness_sorted[:num_parents]



ga_instance = pygad.GA(num_generations= num_generations,
                       sol_per_pop= sol_per_pop,
                       num_parents_mating=2,
                       num_genes= number_of_genes,
                       fitness_func=fitness_func,
                       gene_type=[float, float, float, float,float, float, float, float,float,float],
                       mutation_num_genes=1,
                       gene_space = gene_space,
                       on_generation=on_generation,
                       parent_selection_type=parent_selection_func,
                       on_start= on_start)






#print("Initial Population")
#print(ga_instance.initial_population)

ga_instance.run()



#print("Final Population")
#print(ga_instance.population)



print("______________________________")
solution, solution_fitness, solution_idx = ga_instance.best_solution()
print("Parameters of the best solution : {solution}".format(solution=solution))
print("Fitness value of the best solution = {solution_fitness}".format(solution_fitness=solution_fitness))
print("Index of the best solution : {solution_idx}".format(solution_idx=solution_idx))

# ------------------------------------------------
# Write termination function
t_file = open('C:/Users/TommyHielscher/Desktop/Benchmark Optimization/10 Bar Truss/Control/termination.txt','w')
t_file.write("True")
t_file.close()
# ------------------------------------------------


ga_instance.plot_fitness()


##################Graphinh

""""Listx =  numpy.array(Listx)

#print(Listx)

Listx = numpy.rot90(Listx, k=1, axes=(0, 1))

#print(Listx)

# creating a plot
pixel_plot = plt.figure()



plt.xlabel("X axis label")
plt.ylabel("Y axis label")

# plotting a plot
#pixel_plot.add_axes()

# customizing plot
plt.title("pixel_plot")
pixel_plot = plt.imshow(
    Listx, cmap= 'inferno', interpolation='nearest')

plt.colorbar(pixel_plot)

# save a plot
plt.savefig('pixel_plot.png')

# show plot
plt.show()"""